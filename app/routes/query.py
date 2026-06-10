"""
查询统计模块：多条件查询、分页、ECharts图表数据、Excel导出
"""
import io
from datetime import date, datetime, timedelta

from flask import Blueprint, render_template, request, send_file
from flask_login import login_required

from app import db
from app.models import Credit, OrderRecord, Review, Room, RoomType

qy_bp = Blueprint('query', __name__)


# ==================== 统计辅助函数 ====================
def _get_room_stats():
    total = Room.query.count() or 1
    status = {s: Room.query.filter_by(status=s).count()
              for s in [Room.STATUS_AVAILABLE, Room.STATUS_OCCUPIED,
                        Room.STATUS_RESERVED, Room.STATUS_MAINTENANCE]}
    usage = round((status[Room.STATUS_OCCUPIED] + status[Room.STATUS_RESERVED]) / total * 100, 1)
    return status, usage, total


def _get_revenue_series(today, days):
    result = []
    for i in range(days - 1, -1, -1):
        d = today - timedelta(days=i)
        rev = db.session.query(db.func.sum(OrderRecord.total_fee)).filter(
            db.func.date(OrderRecord.create_time) == d,
            OrderRecord.order_status.in_([OrderRecord.ORDER_CONFIRMED, OrderRecord.ORDER_COMPLETED])
        ).scalar() or 0
        cnt = OrderRecord.query.filter(db.func.date(OrderRecord.create_time) == d).count()
        result.append({'date': d.strftime('%m-%d'), 'revenue': float(rev), 'count': cnt})
    return result


def _get_credit_stats():
    return {
        'total': Credit.query.count(),
        'unpaid_count': Credit.query.filter(Credit.pay_status.in_([Credit.PAY_UNPAID, Credit.PAY_PARTIAL])).count(),
        'total_debt': db.session.query(db.func.sum(Credit.debt_fee)).scalar() or 0,
        'total_unpaid': db.session.query(db.func.sum(Credit.debt_fee)).filter(
            Credit.pay_status.in_([Credit.PAY_UNPAID, Credit.PAY_PARTIAL])).scalar() or 0,
        'total_paid': db.session.query(db.func.sum(Credit.paid_amount)).scalar() or 0,
    }


# ==================== 统计仪表盘 (ECharts) ====================
@qy_bp.route('/stats')
@login_required
def stats_dashboard():
    """数据统计仪表盘"""
    now = datetime.now()
    today = now.date()

    room_status, usage_rate, total_rooms = _get_room_stats()
    today_revenue = db.session.query(db.func.sum(OrderRecord.total_fee)).filter(
        OrderRecord.create_time >= now.replace(hour=0, minute=0, second=0),
        OrderRecord.order_status.in_([OrderRecord.ORDER_CONFIRMED, OrderRecord.ORDER_COMPLETED])
    ).scalar() or 0

    daily_revenue = _get_revenue_series(today, 7)
    monthly_revenue = _get_revenue_series(today, 30)

    type_revenue = db.session.query(
        RoomType.type_name, db.func.sum(OrderRecord.total_fee)
    ).join(Room, Room.type_id == RoomType.id
    ).join(OrderRecord, OrderRecord.room_id == Room.id
    ).filter(OrderRecord.order_status.in_([OrderRecord.ORDER_CONFIRMED, OrderRecord.ORDER_COMPLETED])
    ).group_by(RoomType.type_name).all()

    review_stats = {
        'total': Review.query.count(),
        'avg_rating': round(db.session.query(db.func.avg(Review.rating)).scalar() or 0, 1),
        'dist': {i: Review.query.filter_by(rating=i).count() for i in range(1, 6)}
    }

    return render_template('query/stats.html',
                           room_status=room_status, usage_rate=usage_rate,
                           today_revenue=today_revenue,
                           daily_revenue=daily_revenue,
                           monthly_revenue=monthly_revenue,
                           type_revenue=type_revenue,
                           credit_stats=_get_credit_stats(),
                           review_stats=review_stats,
                           total_rooms=total_rooms)


# ==================== 多条件查询 ====================
@qy_bp.route('/bookings')
@login_required
def query_bookings():
    """预订/入住多条件查询（分页）"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 15, type=int)

    # 筛选参数
    customer = request.args.get('customer', '').strip()
    phone = request.args.get('phone', '').strip()
    status = request.args.get('status', '').strip()
    room_num = request.args.get('room_num', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()

    query = OrderRecord.query
    if customer: query = query.filter(OrderRecord.customer_name.contains(customer))
    if phone: query = query.filter(OrderRecord.phone.contains(phone))
    if status: query = query.filter_by(order_status=status)
    if room_num:
        query = query.join(Room).filter(Room.room_num.contains(room_num))
    if date_from:
        query = query.filter(OrderRecord.create_time >= datetime.strptime(date_from, '%Y-%m-%d'))
    if date_to:
        query = query.filter(OrderRecord.create_time <= datetime.strptime(date_to + ' 23:59:59', '%Y-%m-%d %H:%M:%S'))

    pagination = query.order_by(OrderRecord.create_time.desc()).paginate(
        page=page, per_page=per_page, error_out=False)

    # 汇总
    total_revenue = sum(float(o.total_fee or 0) for o in pagination.items)

    return render_template('query/bookings.html', pagination=pagination,
                           total_revenue=total_revenue,
                           customer=customer, phone=phone, status=status,
                           room_num=room_num, date_from=date_from, date_to=date_to)


@qy_bp.route('/credits')
@login_required
def query_credits():
    """挂账多条件查询（分页）"""
    page = request.args.get('page', 1, type=int)
    company = request.args.get('company', '').strip()
    pay_status = request.args.get('pay_status', '').strip()

    query = Credit.query
    if company: query = query.filter(Credit.company_name.contains(company))
    if pay_status: query = query.filter_by(pay_status=pay_status)

    pagination = query.order_by(Credit.create_time.desc()).paginate(
        page=page, per_page=15, error_out=False)

    total_debt = sum(float(c.debt_fee or 0) for c in pagination.items)
    total_unpaid = sum(float(c.remaining_debt) for c in pagination.items)

    return render_template('query/credits.html', pagination=pagination,
                           total_debt=total_debt, total_unpaid=total_unpaid,
                           company=company, pay_status=pay_status)


# ==================== Excel 导出 ====================
@qy_bp.route('/export/bookings')
@login_required
def export_bookings():
    """导出预订数据为 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = '预订记录'

    # 表头
    headers = ['订单ID', '客户姓名', '电话', '房间号', '入住时间', '退房时间', '费用(元)', '状态', '续费次数', '创建时间']
    header_fill = PatternFill(start_color='0d6efd', end_color='0d6efd', fill_type='solid')
    header_font = Font(name='Arial', color='ffffff', bold=True, size=11)
    body_font = Font(name='Arial', size=10)
    currency_fmt = '#,##0.00'

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    orders = OrderRecord.query.order_by(OrderRecord.create_time.desc()).all()
    for row, o in enumerate(orders, 2):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).font = body_font
        ws.cell(row=row, column=1, value=o.id).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=o.customer_name)
        ws.cell(row=row, column=3, value=o.phone)
        ws.cell(row=row, column=4, value=o.room.room_num if o.room else '')
        ws.cell(row=row, column=5, value=o.check_in.strftime('%Y-%m-%d %H:%M') if o.check_in else '')
        ws.cell(row=row, column=6, value=o.check_out.strftime('%Y-%m-%d %H:%M') if o.check_out else '未退房')
        ws.cell(row=row, column=7, value=float(o.total_fee)).number_format = currency_fmt
        ws.cell(row=row, column=8, value=o.order_status).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=9, value=o.renewal_count).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=10, value=o.create_time.strftime('%Y-%m-%d %H:%M') if o.create_time else '')

    col_widths = [8, 14, 16, 10, 18, 18, 14, 10, 10, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'预订记录_{date.today()}.xlsx')


@qy_bp.route('/export/credits')
@login_required
def export_credits():
    """导出挂账数据为 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = '挂账记录'

    headers = ['挂账ID', '单位名称', '订单ID', '挂账金额', '已还金额', '剩余欠款', '还款状态', '创建时间']
    header_fill = PatternFill(start_color='dc3545', end_color='dc3545', fill_type='solid')
    header_font = Font(name='Arial', color='ffffff', bold=True, size=11)
    body_font = Font(name='Arial', size=10)
    currency_fmt = '#,##0.00'

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    credits = Credit.query.order_by(Credit.create_time.desc()).all()
    for row, c in enumerate(credits, 2):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).font = body_font
        ws.cell(row=row, column=1, value=c.id).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=c.company_name)
        ws.cell(row=row, column=3, value=c.order_id).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=4, value=float(c.debt_fee)).number_format = currency_fmt
        ws.cell(row=row, column=5, value=float(c.paid_amount or 0)).number_format = currency_fmt
        ws.cell(row=row, column=6, value=float(c.remaining_debt)).number_format = currency_fmt
        ws.cell(row=row, column=7, value=c.pay_status).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=8, value=c.create_time.strftime('%Y-%m-%d') if c.create_time else '')

    col_widths = [8, 22, 10, 16, 16, 16, 12, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'挂账记录_{date.today()}.xlsx')
